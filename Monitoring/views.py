from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from .forms import LoginForm, GuestEditForm, UserRegistrationForm, UserEditForm, StudentEditForm, StaffEditForm, TestForm, ProfileForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .models import User, Student, Staff, TelegramProfile, Guest, Test, UserTest, UserAnswer, Faculty, Scale, Choice, Question, Result, BachSpecDirection, MasterDirection, PhdDirection, BachSpecProfile, MasterProfile, PhdProfile
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from openpyxl import Workbook
from django.utils import timezone
from statistics import mode, median, variance, stdev, median, mean
from collections import defaultdict
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import now
from django.utils.dateparse import parse_datetime
import json
import hmac
import hashlib
import logging
from django.contrib import messages
from functools import wraps
from django.http import HttpResponseForbidden
from django.conf import settings


def telegram_login(request):
    if request.method == 'POST':
        logger = logging.getLogger('Monitoring')
        body_unicode = request.body.decode('utf-8')
        request_data = json.loads(body_unicode)
        logger.info(f'Получен POST-запрос на /auth/telegram/: {request_data}')
        # Проверка подписи токена
        data_check_string = "\\n".join([
            f"{key}={value}" for key, value in sorted(request_data.items())
            if key != "hash"
        ])

        secret_key = hmac.new(settings.TELEGRAM_BOT_TOKEN.encode('utf-8'), digestmod=hashlib.sha256).hexdigest()
        signature = hashlib.sha256((data_check_string + secret_key).encode()).hexdigest()

        #if signature != request_data["hash"]:
            #logger.warning(f'Подпись токена неверна: {request_data}')  # Сообщаем о неудачной проверке подписи
            #return JsonResponse({"error": "Invalid token"}, status=400)

        # Получение данных пользователя
        telegram_id = int(request_data["id"])
        telegram_username = request_data["username"] or ""

        logger.info(f'Успешно получили данные пользователя Telegram: {telegram_id}, {telegram_username}')
        # Присоединение профиля к новому пользователю, если его ещё нет
        try:
            # Пытаемся найти пользователя по telegram_id
            user = User.objects.get(username=telegram_username)
            next_url = '/acc/tests/'
        except:
            # Пользователь не найден, создаём нового пользователя
            user, created = User.objects.get_or_create(username=telegram_username)
            next_url = '/edit/'
            profile, _ = TelegramProfile.objects.get_or_create(
            telegram_id=telegram_id,
            defaults={
                "telegram_username": telegram_username,
                "user": user
            }
        )
            student = Student.objects.get_or_create(user=user)            
            logger.info(f'Пользователь {profile.user.username} создан')

        # Авторизация пользователя
        login(request, user)
        request.session['login_method'] = 'telegram' 
        if not user.is_authenticated:
            logger.info(f'Пользователь {user.username} не авторизован.')
        else:
            logger.info(f'Пользователь {user.username} авторизован.')
        return JsonResponse({
        'message': 'Авторизация прошла успешно!',
        'next_url': next_url
    }, status=200)


def admin_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if request.session.get('login_method') != 'password':
            return HttpResponseForbidden("Доступ только по логину и паролю.")
        return view_func(request, *args, **kwargs)
    return _wrapped_view


# Вход 
def user_login(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            user = authenticate(username=cd['username'], password=cd['password'])
            if user is not None:
                if user.is_active:
                    login(request, user)
                    request.session['login_method'] = 'password'
                    return redirect(settings.ADMIN_REDIRECT_URL)
                else:
                    return HttpResponse('Disabled account')
            else:
                return HttpResponse('Invalid login')
    else:
        form = LoginForm()
    return render(request, 'account/login.html', {'form': form})



# Регистрация
def register(request):
    if request.method == 'POST':
        user_form = UserRegistrationForm(request.POST)
        if user_form.is_valid():
            phone_number = user_form.cleaned_data['phone_number']
            new_user = User.objects.create_user(username=phone_number, password=user_form.cleaned_data['password'])
            new_user.save()
            student = Student.objects.create(user=new_user)
            login(request, new_user)
            return redirect('edit')
    else:
        user_form = UserRegistrationForm()
    return render(request, 'account/register.html', 
                  {'user_form': user_form})


def load_directions(request):
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'error': 'Invalid request'}, status=400)
    faculty_id = request.GET.get('faculty')
    if not faculty_id:
        return JsonResponse({'error': 'Faculty ID is required'}, status=400)
    level = request.GET.get('education_level', '')
    try:
        if level in ['бакалавриат', 'специалитет']:
            queryset = BachSpecDirection.objects.filter(faculty_id=faculty_id)
        elif level == 'магистратура':
            queryset = MasterDirection.objects.filter(faculty_id=faculty_id)
        elif level == 'аспирантура':
            queryset = PhdDirection.objects.filter(faculty_id=faculty_id)
        else:
            return JsonResponse({'directions': []})
        directions = list(queryset.order_by('direction_name').values('id', 'direction_name'))
        return JsonResponse({'directions': directions})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def load_profiles(request):
    if request.method != 'GET':
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    try:
        direction_id = request.GET.get('direction')
        education_level = request.GET.get('education_level')
        if not direction_id:
            return JsonResponse({'error': 'Direction ID is required'}, status=400)
        try:
            direction_id = int(direction_id)
        except ValueError:
            return JsonResponse({'error': 'Invalid direction ID'}, status=400)
        if education_level in ['бакалавриат', 'специалитет']:
            profiles_qs = BachSpecProfile.objects.filter(direction_id=direction_id)
        elif education_level == 'магистратура':
            profiles_qs = MasterProfile.objects.filter(direction_id=direction_id)
        elif education_level == 'аспирантура':
            profiles_qs = PhdProfile.objects.filter(direction_id=direction_id)
        else:
            return JsonResponse({'error': 'Invalid education level'}, status=400)
        profiles = []
        for profile in profiles_qs:
            profiles.append({
                'id': profile.id,
                'name': profile.profile,
                'education_form': profile.education_form,
                'education_form_display': profile.get_education_form_display()
            })
        return JsonResponse({'profiles': profiles})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# Персональные данные
def edit(request):
    user = request.user
    if request.method == 'POST':
        role = request.POST.get('role')
        data = request.POST.copy()
        faculty_list = request.POST.getlist('faculty')
        faculty_clean = [v for v in faculty_list if v.strip()]
        if faculty_clean:
            data['faculty'] = faculty_clean[0]
        else:
            data['faculty'] = ''
        user_form = UserEditForm(data, instance=user)

        if role == 'Студент':
            student_instance = getattr(user, 'student', None)
            profile_form = StudentEditForm(data=data, instance=student_instance)
        elif role == 'Сотрудник':
            staff_instance = getattr(user, 'staff', None)
            profile_form = StaffEditForm(data=data, instance=staff_instance)
        elif role == 'Гость':
            guest_instance = getattr(user, 'guest', None)
            profile_form = GuestEditForm(data=data, instance=guest_instance)
        else:
            profile_form = None
        if user_form.is_valid() and (profile_form is None or profile_form.is_valid()):
            user_obj = user_form.save()
            if role == 'Студент':
                Staff.objects.filter(user=user_obj).delete()
                Guest.objects.filter(user=user_obj).delete()
                profile_obj = profile_form.save(commit=False)
                profile_obj.user = user_obj
                profile_obj.role = 'Студент'

                direction_id = profile_obj.direction
                if profile_obj.education_level in ['бакалавриат', 'специалитет']:
                    direction_obj = BachSpecDirection.objects.filter(id=direction_id).first()
                elif profile_obj.education_level == 'магистратура':
                    direction_obj = MasterDirection.objects.filter(id=direction_id).first()
                elif profile_obj.education_level == 'аспирантура':
                    direction_obj = PhdDirection.objects.filter(id=direction_id).first()
                else:
                    direction_obj = None

                if direction_obj:
                    profile_obj.direction = direction_obj.direction_name

                profile_id = profile_obj.profile
                if profile_obj.education_level in ['бакалавриат', 'специалитет']:
                    profile_instance = BachSpecProfile.objects.filter(id=profile_id).first()
                elif profile_obj.education_level == 'магистратура':
                    profile_instance = MasterProfile.objects.filter(id=profile_id).first()
                elif profile_obj.education_level == 'аспирантура':
                    profile_instance = PhdProfile.objects.filter(id=profile_id).first()
                else:
                    profile_instance = None
                if profile_instance:
                    profile_obj.profile = profile_instance.profile
                    profile_obj.study_form = profile_instance.education_form

                profile_obj.save()
                messages.success(request, 'Данные студента сохранены!')

            elif role == 'Сотрудник':
                Student.objects.filter(user=user_obj).delete()
                Guest.objects.filter(user=user_obj).delete()
                profile_obj = profile_form.save(commit=False)
                profile_obj.user = user_obj
                profile_obj.role = 'Сотрудник'
                profile_obj.save()
                messages.success(request, 'Данные сотрудника сохранены!')

            elif role == 'Гость':
                Student.objects.filter(user=user_obj).delete()
                Staff.objects.filter(user=user_obj).delete()
                profile_obj = profile_form.save(commit=False)
                profile_obj.user = user_obj
                profile_obj.role = 'Гость'
                profile_obj.save()
                messages.success(request, 'Данные гостя сохранены!')
            return redirect('user_login')

        else:
            messages.error(request, 'Проверьте правильность данных в форме.')

    else:
        profile_form = None
        role = ''
        if hasattr(user, 'student'):
            role = 'Студент'
            student_instance = getattr(user, 'student', None)
            profile_form = StudentEditForm(instance=student_instance)
        elif hasattr(user, 'staff'):
            role = 'Сотрудник'
            staff_instance = getattr(user, 'staff', None)
            profile_form = StaffEditForm(instance=staff_instance)
        elif hasattr(user, 'guest'):
            role = 'Гость'
            guest_instance = getattr(user, 'guest', None)
            profile_form = GuestEditForm(instance=guest_instance)
        user_form = UserEditForm(instance=user)
    return render(
        request,
        'registration/edit.html',
        {
            'user_form': user_form,
            'profile_form': profile_form,
            'role': role,
        }
    )
    

@login_required(login_url='login')
# Прохождение теста и подсчёт баллов
def start_test(request, test_id, question_index=1):
    test = get_object_or_404(Test, id=test_id)
    questions = test.questions.all()
    questions_list = list(questions.values_list('id', flat=True))
    questions_json = json.dumps(questions_list)
    user_test_id = request.session.get('user_test_id')
    user_test = None
    if user_test_id:
        try:
            user_test = UserTest.objects.get(id=user_test_id)
            if user_test.test.id != test.id:
                user_test = None
                request.session.pop('user_test_id')
                request.session.pop('start_time', None)
        except UserTest.DoesNotExist:
            user_test = None
            request.session.pop('user_test_id')
            request.session.pop('start_time', None)
    if user_test is None and question_index == 1:
        user_test = UserTest(user=request.user, test=test)
        user_test.save()
        request.session['user_test_id'] = user_test.id
        request.session['start_time'] = now().isoformat()
    elif user_test is None:
        return redirect('start_test', test_id=test_id, question_index=1)
    start_time_str = request.session.get('start_time')
    try:
        start_dt = parse_datetime(start_time_str)
    except Exception:
        start_dt = now()
    try:
        question_index = int(question_index)
    except (TypeError, ValueError):
        question_index = 1
    if question_index < 1:
        question_index = 1
    if question_index > questions.count():
        question_index = questions.count()
    question = questions[question_index - 1]
    scale = question.scales.first()
    if scale:
        min_s = scale.min_score
        max_s = scale.max_score
        if scale.type == "default":
            max_range = list(range(min_s, max_s + 1))
            minus = []
            twoside = []
        elif scale.type == "twoside":
            count = max_s - min_s + 1
            center = (count - 1) // 2
            twoside = [(i - center) for i in range(count)]
            max_range = []
            minus = []
        elif scale.type == "minus":
            count = max_s - min_s + 1
            half = count // 2
            minus = list(range(-half, half + 1))
            max_range = []
            twoside = []
        if scale and scale.type in ["default", "twoside"] and scale.inverse:
            if scale.type == "default":
                max_range = list(reversed(max_range))
            elif scale.type == "twoside":
                twoside = list(reversed(twoside))
    else:
        max_range = []
        minus = []
        twoside = []
    selected_answer = None
    user_answer = UserAnswer.objects.filter(user_test=user_test, question=question).first()
    if user_answer:
        if question.question_type == 'choice':
            selected_answer = user_answer.chosen_choice.id
        else:
            if scale and scale.type == 'twoside':
                score = user_answer.score
                idx = score - scale.min_score
                center = (scale.max_score - scale.min_score + 1 - 1) // 2
                display_value = idx - center
                selected_answer = display_value
            else:
                selected_answer = user_answer.score
    error_message = None
    test_time = test.time
    duration_sec = test_time * 60 if test_time else None
    time_is_up = False
    if duration_sec is not None and (now() - start_dt).total_seconds() > duration_sec:
        time_is_up = True
    if time_is_up or request.POST.get('_force_finish'): 
        for q in questions:
            if not UserAnswer.objects.filter(user_test=user_test, question=q).exists():
                UserAnswer.objects.create(user_test=user_test, question=q, score=0)
        del request.session['user_test_id']
        if 'start_time' in request.session:
            del request.session['start_time']
        return redirect('test_results', user_test_id=user_test.id)
    if request.method == 'POST':
        all_post_keys = list(request.POST.keys())
        answer_key = str(question.id) if str(question.id) in all_post_keys else None
        if answer_key:
            chosen_value = request.POST.get(answer_key)
            chosen_choice_obj = None
            new_score = 0
            if question.question_type == 'choice':
                chosen_choice_obj = question.choices.filter(id=chosen_value).first()
                if chosen_choice_obj:
                    new_score = chosen_choice_obj.score
            else:
                if scale:
                    min_s = scale.min_score
                    max_s = scale.max_score
                    if scale.type == "default":
                        if chosen_value and str(chosen_value).lstrip('-').isdigit():
                            new_score = int(chosen_value)
                        else:
                            new_score = 0
                    elif scale.type == "twoside":
                        count = max_s - min_s + 1
                        center = (count - 1) // 2
                        if chosen_value is not None and (str(chosen_value).lstrip('-').isdigit() or chosen_value == '0'):
                            display_value = int(chosen_value)
                            idx = display_value + center
                            if 0 <= idx < count:
                                new_score = min_s + idx
                            else:
                                new_score = min_s
                        else:
                            new_score = min_s
                    elif scale.type == "minus":
                        if chosen_value is not None and (str(chosen_value).lstrip('-').isdigit() or chosen_value == '0'):
                            display_value = int(chosen_value)
                            new_score = display_value
                        else:
                            new_score = 0
                    if scale and scale.type in ["default", "twoside"] and scale.inverse:
                        if scale.type == "default":
                            max_range = list(range(min_s, max_s + 1))
                        elif scale.type == "twoside":
                            count = max_s - min_s + 1
                            center = (count - 1) // 2
                            twoside = [(i - center) for i in range(count)]
                else:
                    new_score = 0
            if user_answer:
                if (user_answer.chosen_choice and user_answer.chosen_choice.id != (chosen_choice_obj.id if chosen_choice_obj else None)) \
                    or (not user_answer.chosen_choice and user_answer.score != new_score):
                    if question.question_type == 'choice':
                        user_answer.chosen_choice = chosen_choice_obj
                        user_answer.score = new_score
                    else:
                        user_answer.chosen_choice = None
                        user_answer.score = new_score
                    user_answer.save()
            else:
                UserAnswer.objects.create(
                    user_test=user_test,
                    question=question,
                    chosen_choice=chosen_choice_obj,
                    score=new_score
                )
            selected_answer = chosen_choice_obj.id if chosen_choice_obj else new_score
            user_test.result = sum(UserAnswer.objects.filter(user_test=user_test).values_list('score', flat=True))
            user_test.save()
            if question_index < questions.count():
                return redirect('start_test', test_id=test_id, question_index=question_index + 1)
            else:
                return redirect('start_test', test_id=test_id, question_index=question_index)
        if 'back' in request.POST and question_index > 1:
            return redirect('start_test', test_id=test_id, question_index=question_index - 1)
        if 'close' in request.POST:
            unanswered = [q for q in questions if q.is_mandatory and not UserAnswer.objects.filter(user_test=user_test, question=q).exists()]
            if unanswered:
                error_message = "Вы не ответили на все обязательные вопросы. Завершение теста невозможно."
            else:
                del request.session['user_test_id']
                if 'start_time' in request.session:
                    del request.session['start_time']
                return redirect('test_results', user_test_id=user_test.id)
        if 'next' in request.POST:
            if question_index >= len(questions):
                del request.session['user_test_id']
                return redirect('test_results', user_test_id=user_test.id)
            return redirect('start_test', test_id=test_id, question_index=question_index + 1)
    answered_question_ids = list(
        UserAnswer.objects.filter(user_test=user_test)
        .values_list('question_id', flat=True)
    )
    return render(request, 'tests/start_test.html', {
        'question': question,
        'test': test,
        'question_index': question_index,
        'scale': scale,
        'max_range': max_range,
        'minus': minus,
        'twoside': twoside,
        'questions_count': questions.count(),
        'selected_answer': selected_answer,
        'error_message': error_message,
        'test_time': test.time,
        'start_time': start_dt,
        'questions_list': questions_json,
        'user_test_id': user_test.id,
        'answered_question_ids': answered_question_ids,
    })


@login_required(login_url='login')
# Вывод результата теста
def test_results(request, user_test_id):
    user_test = get_object_or_404(UserTest, id=user_test_id)
    answers = user_test.user_answers.all()
    max_score = 0
    min_score = 0
    for question in user_test.test.questions.all():
        q_max = None
        q_min = None
        if question.choices.exists():
            choices = question.choices.all()
            q_max = max((c.score for c in choices if c.score is not None), default=0)
            q_min = min((c.score for c in choices if c.score is not None), default=0)
        elif question.scales.exists():
            scales = question.scales.all()
            q_max = max((s.max_score for s in scales if s.max_score is not None), default=0)
            q_min = min((s.min_score for s in scales if s.min_score is not None), default=0)
        max_score += q_max if q_max is not None else 0
        min_score += q_min if q_min is not None else 0
    percentage = 0
    if max_score != min_score:
        raw_percent = ((user_test.result - min_score) / (max_score - min_score)) * 100
        percentage = max(0, min(100, round(raw_percent, 1)))
    else:
        percentage = 0
    result = None
    results = Result.objects.filter(test=user_test.test)
    for res in results:
        if res.condition in 'greater':
            if user_test.result > res.min_score:
                result = res
                break 
        elif res.condition == 'greater_or_equal':
            if user_test.result >= res.min_score:
                result = res
                break 
        elif res.condition in 'less':
            if user_test.result < res.min_score:
                result = res
                break
        elif res.condition == 'less_or_equal':
            if user_test.result <= res.min_score:
                result = res
                break 
        elif res.condition in 'between':
            if res.min_score <= user_test.result <= res.max_score:
                result = res
                break
        elif res.condition in 'between_exclusive':
            if res.min_score < user_test.result < res.max_score:
                result = res
                break
        elif res.condition in 'between_inclusive':
            if res.min_score <= user_test.result < res.max_score:
                result = res
                break
        elif res.condition in 'between_matching':
            if res.min_score < user_test.result <= res.max_score:
                result = res
                break
        elif res.condition == 'equal':
            if user_test.result == res.min_score:
                result = res
                break
    if not result and results.exists():
        result = results.order_by('min_score').first()
    if result:
        user_test.result_description = result.result_description
        user_test.additional_info = result.additional_info
        user_test.result_image = result.image
    user_test.completed_at = timezone.now()
    user_test.save()
    user_tests = UserTest.objects.filter(user=request.user).order_by('-completed_at')
    return render(request, 'tests/test_results.html', {
        'user_test': user_test,
        'answers': answers,
        'result': result,
        'user_tests': user_tests,
        'result_image': result.image if result else None,   
        'min_score': min_score,
        'max_score': max_score,
        'percentage': percentage,
    })


@login_required(login_url='login')
# Список тестов
def test_list(request):
    now = timezone.now()
    user = request.user
    tests = Test.objects.filter(start_datetime__lte=now)
    tests = [test for test in tests if test.is_available_for_user(user)]
    tests_count = len(tests)
    return render(request, 'tests/test_list.html', {
        'tests': tests,
        'tests_count': tests_count,
        'now': now,
    })


@login_required(login_url='login')
# Описание теста перед началом тестирования
def test_intro(request, test_id):
    test = get_object_or_404(Test, id=test_id)
    questions_count = test.questions.count()
    return render(request, 'tests/test_intro.html', {
        'test': test,
        'questions_count': questions_count,
        'test_time': test.time, 
    })


@login_required(login_url='login')
# Вывод всех результатов тестирования пользователя
def user_results(request, test_id):
    test = get_object_or_404(Test, id=test_id)
    user_tests = UserTest.objects.filter(user=request.user, test=test).order_by('-completed_at')
    return render(request, 'tests/user_results.html', {
        'user_tests': user_tests,
        'test': test,
    })


@admin_required
# Страница с таблицей пользователей
def admin_users(request):
    search_query = request.GET.get('q', '')
    dob_start = request.GET.get('dob_start', '')
    dob_end = request.GET.get('dob_end', '')
    selected_test = request.GET.get('test_filter', '')
    faculties_selected = request.GET.getlist('faculty')
    roles = request.GET.getlist('roles')
    directions_selected = request.GET.getlist('direction')
    profiles_selected = request.GET.getlist('profile')
    education_levels_selected = request.GET.getlist('education_level')
    selected_grades = request.GET.getlist('grade')
    status = request.GET.get('status', '')
    status_min = request.GET.get('status_min', '')
    status_max = request.GET.get('status_max', '')
    test_date_start = request.GET.get('test_date_start', '')
    test_date_end = request.GET.get('test_date_end', '')
    tests = Test.objects.all()
    students = Student.objects.select_related('user')
    staffs = Staff.objects.select_related('user')
    guests = Guest.objects.select_related('user')

    # Фильтрация по поиску
    if search_query:
        students = students.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(date_of_birth__icontains=search_query) |
            Q(role__icontains=search_query) |
            Q(faculty__icontains=search_query)
        )
        staffs = staffs.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(date_of_birth__icontains=search_query) |
            Q(role__icontains=search_query) |
            Q(faculty__icontains=search_query)
        )
        guests = guests.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(date_of_birth__icontains=search_query) |
            Q(role__icontains=search_query)
        )

    # Фильтрация по дате рождения
    if dob_start:
        students = students.filter(date_of_birth__gte=dob_start)
        staffs = staffs.filter(date_of_birth__gte=dob_start)
        guests = guests.filter(date_of_birth__gte=dob_start)
    if dob_end:
        students = students.filter(date_of_birth__lte=dob_end)
        staffs = staffs.filter(date_of_birth__lte=dob_end)
        guests = guests.filter(date_of_birth__lte=dob_end)

    # фильтры (роль/статус)
    if roles:
        students = students.filter(role__in=roles)
        staffs = staffs.filter(role__in=roles)
        guests = guests.filter(role__in=roles)

    # Фильтрация по факультету, группе и курсу
    if faculties_selected:
        students = students.filter(faculty__in=faculties_selected)
        staffs = staffs.filter(faculty__in=faculties_selected)
        guests = guests.none()
    elif faculties_selected and (directions_selected or selected_grades):
        students = students.filter(faculty__in=faculties_selected)
        if directions_selected:
            students = students.filter(direction__in=directions_selected)
            staffs = staffs.none()
            guests = guests.none()
        if selected_grades:
            students = students.filter(grade__in=selected_grades)
            staffs = staffs.none()
            guests = guests.none()
        staffs = staffs.none()
    elif not faculties_selected and (directions_selected or selected_grades):
        if directions_selected:
            students = students.filter(direction__in=directions_selected)
            staffs = staffs.none()
            guests = guests.none()
        if selected_grades:
            students = students.filter(grade__in=selected_grades)
            staffs = staffs.none()
            guests = guests.none()
        staffs = staffs.none()

    if directions_selected:
        students = students.filter(direction__in=directions_selected)
        staffs = staffs.none()
        guests = guests.none()

    if profiles_selected:
        students = students.filter(profile__in=profiles_selected)
        staffs = staffs.none()
        guests = guests.none()

    if selected_grades:
        students = students.filter(grade__in=selected_grades)
        staffs = staffs.none()
        guests = guests.none()

    if education_levels_selected:
        students = students.filter(education_level__in=education_levels_selected)
        staffs = staffs.none()
        guests = guests.none()
    users = students.union(staffs)
    
    result_scores = set()
    if selected_test and selected_test != '__none__':
        all_attempts = UserTest.objects.filter(test_id=selected_test).order_by('user', '-completed_at')
        latest_attempts_dict = {}
        for attempt in all_attempts:
            if attempt.user_id not in latest_attempts_dict:
                latest_attempts_dict[attempt.user_id] = attempt
        for ut in latest_attempts_dict.values():
            result_scores.add(ut.result)
    else:
        all_attempts = UserTest.objects.all().order_by('user', '-completed_at')
        latest_attempts_dict = {}
        for attempt in all_attempts:
            if attempt.user_id not in latest_attempts_dict:
                latest_attempts_dict[attempt.user_id] = attempt
        for ut in latest_attempts_dict.values():
            result_scores.add(ut.result)
    result_scores = sorted(result_scores)

    def pass_test_date_and_status_filters(user_object, last_test, status_min, status_max, test_date_start, test_date_end):
        if test_date_start:
            if not last_test or not last_test.completed_at or last_test.completed_at.date() < datetime.strptime(test_date_start, "%Y-%m-%d").date():
                return False
        if test_date_end:
            if not last_test or not last_test.completed_at or last_test.completed_at.date() > datetime.strptime(test_date_end, "%Y-%m-%d").date():
                return False
        if status_min or status_max:
            result_value = None
            if last_test and hasattr(last_test, 'result'):
                try:
                    result_value = float(last_test.result)
                except (ValueError, TypeError):
                    return False
                if status_min:
                    try:
                        if result_value < float(status_min):
                            return False
                    except ValueError:
                        pass
                if status_max:
                    try:
                        if result_value > float(status_max):
                            return False
                    except ValueError:
                        pass
            else:
                return False
        return True
    
    combined_list = []
    if selected_test and selected_test != '' and selected_test != '__none__':
        for student in students:
            user_test = student.user.usertest_set.filter(test_id=selected_test).order_by('-completed_at').first()
            if user_test:
                if pass_test_date_and_status_filters(student, user_test, status_min, status_max, test_date_start, test_date_end):
                    student.user_type = 'student'
                    student.last_test = user_test
                    combined_list.append(student)
        for staff in staffs:
            user_test = staff.user.usertest_set.filter(test_id=selected_test).order_by('-completed_at').first()
            if user_test:
                if pass_test_date_and_status_filters(staff, user_test,  status_min, status_max, test_date_start, test_date_end):
                    staff.user_type = 'staff'
                    staff.last_test = user_test
                    combined_list.append(staff)
        for guest in guests:
            user_test = guest.user.usertest_set.filter(test_id=selected_test).order_by('-completed_at').first()
            if user_test:
                if pass_test_date_and_status_filters(guest, user_test,  status_min, status_max, test_date_start, test_date_end):
                    guest.user_type = 'guest'
                    guest.last_test = user_test
                    combined_list.append(guest)

    elif selected_test == '__none__':
        for student in students:
            if not student.user.usertest_set.exists():
                if not (test_date_start or test_date_end or status):
                    student.user_type = 'student'
                    student.last_test = None
                    combined_list.append(student)
        for staff in staffs:
            if not staff.user.usertest_set.exists():
                if not (test_date_start or test_date_end or status):
                    staff.user_type = 'staff'
                    staff.last_test = None
                    combined_list.append(staff)
        for guest in guests:
            if not guest.user.usertest_set.exists():
                if not (test_date_start or test_date_end or status):
                    guest.user_type = 'guest'
                    guest.last_test = None
                    combined_list.append(guest)
    else:
        for student in students:
            user_test = student.user.usertest_set.order_by('-completed_at').first()
            if pass_test_date_and_status_filters(student, user_test, status_min, status_max, test_date_start, test_date_end):
                student.user_type = 'student'
                student.last_test = user_test
                combined_list.append(student)
        for staff in staffs:
            user_test = staff.user.usertest_set.order_by('-completed_at').first()
            if pass_test_date_and_status_filters(staff, user_test, status_min, status_max, test_date_start, test_date_end):
                staff.user_type = 'staff'
                staff.last_test = user_test
                combined_list.append(staff)
        for guest in guests:
            user_test = guest.user.usertest_set.order_by('-completed_at').first()
            if pass_test_date_and_status_filters(guest, user_test, status_min, status_max, test_date_start, test_date_end):
                guest.user_type = 'guest'
                guest.last_test = user_test
                combined_list.append(guest)

    combined_list.sort(key=lambda x: x.user.last_name)
    paginator = Paginator(combined_list, 20)
    page = request.GET.get('page')

    try:
        users = paginator.page(page)
    except PageNotAnInteger:
        users = paginator.page(1)
    except EmptyPage:
        users = paginator.page(paginator.num_pages)

    all_students = Student.objects.all()
    profiles = sorted(set(filter(None, all_students.values_list('profile', flat=True))))
    education_levels = sorted(set(filter(None, all_students.values_list('education_level', flat=True))))
    all_staffs = Staff.objects.all()
    all_guests = Guest.objects.all()
    faculties_qs = all_students.values_list('faculty', flat=True).distinct().union(
        all_staffs.values_list('faculty', flat=True).distinct()
    )
    faculties = sorted(set(filter(None, faculties_qs)))
    grades = sorted(set(filter(None, all_students.values_list('grade', flat=True))))
    directions = sorted(set(filter(None, all_students.values_list('direction', flat=True))))

    return render(
        request,
        'admin/users.html',
        {
            'users': users,
            'search_query': search_query,
            'dob_start': dob_start,
            'dob_end': dob_end,
            'tests': tests,
            'selected_test': selected_test,
            'roles': roles,
            'faculty': faculties_selected,
            'directions_selected': directions_selected, 
            'directions': directions, 
            'selected_grades': selected_grades,
            'faculties': faculties,
            'profiles_selected': profiles_selected,
            'education_level_selected': education_levels_selected,
            'profiles': profiles,
            'education_levels': education_levels,
            'grades': grades,
            'status': status,
            'status_min': status_min,
            'status_max': status_max,
            'test_date_start': test_date_start,
            'test_date_end': test_date_end,
            'status_options': result_scores
        }
    )


def score_in_range(user_test, status_min, status_max):
    try:
        score = float(user_test.result)
    except (ValueError, TypeError, AttributeError):
        return False
    if status_min and score < float(status_min):
        return False
    if status_max and score > float(status_max):
        return False
    return True


@admin_required
# Выгрузка таблицы пользователей в файл
def export_to_excel(request):
    search_query = request.GET.get('q', '')
    dob_start = request.GET.get('dob_start', '')
    dob_end = request.GET.get('dob_end', '')
    selected_test = request.GET.get('test_filter', '')
    faculties_selected = request.GET.getlist('faculty')
    roles = request.GET.getlist('roles')
    directions_selected = request.GET.getlist('direction')
    profiles_selected = request.GET.getlist('profile')
    education_levels_selected = request.GET.getlist('education_level')
    selected_grades = request.GET.getlist('grade')
    status = request.GET.get('status', '')
    status_min = request.GET.get('status_min', '')
    status_max = request.GET.get('status_max', '')
    test_date_start = request.GET.get('test_date_start', '')
    test_date_end = request.GET.get('test_date_end', '')
    students = Student.objects.select_related('user').all()
    staffs = Staff.objects.select_related('user').all()
    guests = Guest.objects.select_related('user').all()

    # Поиск
    if search_query:
        students = students.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(date_of_birth__icontains=search_query) |
            Q(role__icontains=search_query) |
            Q(faculty__icontains=search_query)
        )
        staffs = staffs.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(date_of_birth__icontains=search_query) |
            Q(role__icontains=search_query) |
            Q(faculty__icontains=search_query)
        )
        guests = guests.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(date_of_birth__icontains=search_query) |
            Q(role__icontains=search_query)
        )
    # Фильтр по дате рождения
    if dob_start:
        students = students.filter(date_of_birth__gte=dob_start)
        staffs = staffs.filter(date_of_birth__gte=dob_start)
        guests = guests.filter(date_of_birth__gte=dob_start)
    if dob_end:
        students = students.filter(date_of_birth__lte=dob_end)
        staffs = staffs.filter(date_of_birth__lte=dob_end)
        guests = guests.filter(date_of_birth__lte=dob_end)

    # Фильтр по роли
    if roles:
        students = students.filter(role__in=roles)
        staffs = staffs.filter(role__in=roles)
        guests = guests.filter(role__in=roles)

    if faculties_selected:
        students = students.filter(faculty__in=faculties_selected)
        staffs = staffs.filter(faculty__in=faculties_selected)
        guests = guests.none()
    elif faculties_selected and (directions_selected or selected_grades):
        students = students.filter(faculty__in=faculties_selected)
        if directions_selected:
            students = students.filter(direction__in=directions_selected)
            guests = guests.none()
            staffs = staffs.none()
        if selected_grades:
            students = students.filter(grade__in=selected_grades)
            guests = guests.none()
            staffs = staffs.none()
        staffs = staffs.none()
    elif not faculties_selected and (directions_selected or selected_grades):
        if directions_selected:
            students = students.filter(direction__in=directions_selected)
            guests = guests.none()
            staffs = staffs.none()
        if selected_grades:
            students = students.filter(grade__in=selected_grades)
            guests = guests.none()
            staffs = staffs.none()
        staffs = staffs.none()

    if directions_selected:
        students = students.filter(direction__in=directions_selected)
        guests = guests.none()
        staffs = staffs.none()

    combined_list = []
    if selected_test and selected_test != '' and selected_test != '__none__':
        for student in students:
            user_test = student.user.usertest_set.filter(test_id=selected_test).order_by('-completed_at').first()
            if user_test and score_in_range(user_test, status_min, status_max) and pass_test_date_and_status_filters(student, user_test, status, test_date_start, test_date_end):
                student.user_type = 'student'
                student.last_test = user_test
                combined_list.append(student)
        for staff in staffs:
            user_test = staff.user.usertest_set.filter(test_id=selected_test).order_by('-completed_at').first()
            if user_test and score_in_range(user_test, status_min, status_max) and pass_test_date_and_status_filters(staff, user_test, status, test_date_start, test_date_end):
                staff.user_type = 'staff'
                staff.last_test = user_test
                combined_list.append(staff)
        for guest in guests:
            user_test = guest.user.usertest_set.filter(test_id=selected_test).order_by('-completed_at').first()
            if user_test and score_in_range(user_test, status_min, status_max) and pass_test_date_and_status_filters(guest, user_test, status, test_date_start, test_date_end):
                guest.user_type = 'guest'
                guest.last_test = user_test
                combined_list.append(guest)
    elif selected_test == '__none__':
        for student in students:
            if not student.user.usertest_set.exists():
                if not (test_date_start or test_date_end or status):
                    student.user_type = 'student'
                    student.last_test = None
                    combined_list.append(student)
        for staff in staffs:
            if not staff.user.usertest_set.exists():
                if not (test_date_start or test_date_end or status):
                    staff.user_type = 'staff'
                    staff.last_test = None
                    combined_list.append(staff)
        for guest in guests:
            if not guest.user.usertest_set.exists():
                if not (test_date_start or test_date_end or status):
                    guest.user_type = 'guest'
                    guest.last_test = None
                    combined_list.append(guest)
    else:
        for student in students:
            user_test = student.user.usertest_set.order_by('-completed_at').first()
            if pass_test_date_and_status_filters(student, user_test, status, test_date_start, test_date_end):
                student.user_type = 'student'
                student.last_test = user_test
                combined_list.append(student)
        for staff in staffs:
            user_test = staff.user.usertest_set.order_by('-completed_at').first()
            if pass_test_date_and_status_filters(staff, user_test, status, test_date_start, test_date_end):
                staff.user_type = 'staff'
                staff.last_test = user_test
                combined_list.append(staff)
        for guest in guests:
            user_test = guest.user.usertest_set.order_by('-completed_at').first()
            if pass_test_date_and_status_filters(guest, user_test, status, test_date_start, test_date_end):
                guest.user_type = 'guest'
                guest.last_test = user_test
                combined_list.append(guest)

    combined_list.sort(key=lambda x: x.user.last_name)
    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"

    basic_headers = [
        'id', 'Фамилия', 'Имя', 'Отчество', 'Дата рождения', 'Роль',
        'Факультет', 'Направление', 'Профиль', 'Уровень образования', 'Курс', 'Дата тестирования', 'Время тестирования'
    ]
    headers = list(basic_headers)
    question_list = []
    test_title = ""
    with_ball_columns = False

    if selected_test and selected_test not in ['', '__none__']:
        with_ball_columns = True
        try:
            test_obj = Test.objects.get(id=selected_test)
            test_title = test_obj.title
            question_list = list(test_obj.questions.all().order_by('id'))
        except Test.DoesNotExist:
            test_title = "Не найден"
            question_list = []
        for i in range(1, len(question_list) + 1):
            headers.append(f'Вопрос {i} (баллы)')
        headers.append('Текущее состояние')
    ws.append(headers)

    for user in combined_list:
        faculty = getattr(user, 'faculty', '-') or '-'
        direction = getattr(user, 'direction', '-') if getattr(user, 'user_type', '') == 'student' else '-'
        profile = getattr(user, 'profile', '-') if getattr(user, 'user_type', '') == 'student' else '-'
        education_level = getattr(user, 'education_level', '-') if getattr(user, 'user_type', '') == 'student' else '-'
        grade = getattr(user, 'grade', '-') if getattr(user, 'user_type', '') == 'student' else '-'
        
        test_date = "нет даты"
        test_time = "нет времени"
        if user.last_test and user.last_test.completed_at:
            test_date = user.last_test.completed_at.strftime("%d.%m.%Y")
            test_time = user.last_test.completed_at.strftime("%H:%M")

        row = [
            user.id,
            user.user.last_name,
            user.user.first_name,
            getattr(user, 'patronymic', ''),
            user.date_of_birth.strftime("%d.%m.%Y") if user.date_of_birth else '',
            user.role,
            faculty,
            direction,
            profile,
            education_level,
            grade,
            test_date,
            test_time 
        ]
        if with_ball_columns:
            qid2score = {}
            if user.last_test and question_list:
                answers = UserAnswer.objects.filter(user_test=user.last_test)
                for ans in answers:
                    qid2score[ans.question_id] = ans.score
            for question in question_list:
                score = qid2score.get(question.id, "Нет ответа")
                row.append(score)
            row.append(user.last_test.result if user.last_test else "Нет результата")
        ws.append(row)

    if with_ball_columns:
        row_index = 1
        col_index = len(headers) + 3
        ws.cell(row=row_index, column=col_index, value="Тест:")
        ws.cell(row=row_index+1, column=col_index, value=test_title)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=users.xlsx'
    wb.save(response)
    return response


@admin_required
# Редактирование данных о пользователях в таблице
def bulk_edit_users(request):
    if request.method == 'POST':
        role_map = {
            "student": "Студент",
            "staff": "Сотрудник", 
            "guest": "Гость"
        }
        role_map_reverse = {
            "Студент": "student",
            "Сотрудник": "staff",
            "Гость": "guest",
            "student": "student",
            "staff": "staff",
            "guest": "guest"
        }
        try:
            data = json.loads(request.body)
        except Exception as e:
            return JsonResponse({'status': 'bad_json'}, status=400)
        response_data = {'status': 'ok', 'details': []}
        for edit in data.get('edits', []):
            user_id = edit.get('user_id')
            fields = edit.get('fields', {})
            try:
                user_obj = User.objects.get(id=user_id)
            except User.DoesNotExist:
                continue
            current_role = None
            if hasattr(user_obj, 'student'):
                current_role = 'student'
            elif hasattr(user_obj, 'staff'):
                current_role = 'staff'
            elif hasattr(user_obj, 'guest'):
                current_role = 'guest'
                
            new_role_value = fields.get('role', current_role)
            new_role = role_map_reverse.get(new_role_value, current_role)
            if new_role not in ['student', 'staff', 'guest']:
                new_role = current_role
                
            if 'last_name' in fields:
                user_obj.last_name = fields['last_name']
            if 'first_name' in fields:
                user_obj.first_name = fields['first_name']
            user_obj.save()
            if new_role == current_role and current_role:
                profile = getattr(user_obj, current_role)
                for field_name, field_value in fields.items():
                    if field_name in ['last_name', 'first_name', 'role']:
                        continue
                    if hasattr(profile, field_name):
                        if field_name == 'date_of_birth':
                            if field_value:
                                try:
                                    setattr(profile, field_name, datetime.strptime(field_value, '%Y-%m-%d').date())
                                except:
                                    profile.date_of_birth = None
                            else:
                                profile.date_of_birth = None
                        else:
                            setattr(profile, field_name, field_value if field_value is not None else "")
                profile.role = role_map.get(new_role, "Гость")
                profile.save()
            else:
                if current_role and hasattr(user_obj, current_role):
                    getattr(user_obj, current_role).delete()
                profile_data = {
                    'user': user_obj,
                    'date_of_birth': None,
                    'patronymic': fields.get('patronymic', ''),
                    'role': role_map.get(new_role, "Гость")
                }
                if 'date_of_birth' in fields and fields['date_of_birth']:
                    try:
                        profile_data['date_of_birth'] = datetime.strptime(fields['date_of_birth'], '%Y-%m-%d').date()
                    except:
                        pass
                if new_role == 'student':
                    profile_data.update({
                        'faculty': fields.get('faculty', ''),
                        'direction': fields.get('direction', ''),
                        'grade': fields.get('grade', ''),
                        'profile': fields.get('profile', ''),
                        'education_level': fields.get('education_level', ''),
                    })
                    Student.objects.create(**profile_data)
                elif new_role == 'staff':
                    profile_data.update({
                        'faculty': fields.get('faculty', ''),
                    })
                    Staff.objects.create(**profile_data)
                elif new_role == 'guest':
                    Guest.objects.create(**profile_data)
        return JsonResponse(response_data)
    return JsonResponse({'status': 'error'}, status=400)


@admin_required
# Профиль конкретного пользователя
def user_profile(request, user_id):
    user = get_object_or_404(User, id=user_id)
    student_profile = getattr(user, 'student', None)
    staff_profile = getattr(user, 'staff', None)
    guest_profile = getattr(user, 'guest', None)
    user_tests = user.usertest_set.all().order_by('-completed_at')
    all_tests = Test.objects.all()
    selected_test_results = None
    selected_test = None
    if request.method == 'POST':
        selected_test = request.POST.get('selected_test')
        if selected_test:
            selected_test_results = UserTest.objects.filter(user=user, test_id=selected_test).first()
    else:
        selected_test = None
    if selected_test_results:
        user_tests = UserTest.objects.filter(user=user, test=selected_test_results.test).order_by('-completed_at')
    else:
        user_tests = user.usertest_set.all().order_by('-completed_at')
    context = {
        'user': user,
        'student_profile': student_profile,
        'staff_profile': staff_profile,
        'guest_profile': guest_profile,
        'user_tests': user_tests,
        'all_tests': all_tests,
        'selected_test_results': selected_test_results,
        'selected_test': selected_test,
        'answers': selected_test_results.user_answers.all() if selected_test_results else [],
    }
    return render(request, 'admin/user_profile.html', context)


@admin_required
# Редактирование профиля конкретного пользователя
def edit_user_profile(request, user_id):
    user = get_object_or_404(User, id=user_id)
    initial = {
        'first_name': user.first_name,
        'last_name': user.last_name,
        'patronymic': '',
        'date_of_birth': None,
        'role': 'гость',
        'faculty': '',
        'direction': '',
        'grade': '',
        'profile': '',
        'study_form': '',
        'education_level': '',
    }
    if hasattr(user, 'student'):
        profile = user.student
        initial.update({
            'patronymic': profile.patronymic,
            'date_of_birth': profile.date_of_birth,
            'role': 'студент',
            'faculty': profile.faculty,
            'direction': profile.direction, 
            'grade': profile.grade,
            'profile': profile.profile,
            'study_form': profile.study_form,
            'education_level': profile.education_level,
        })
    elif hasattr(user, 'staff'):
        profile = user.staff
        initial.update({
            'patronymic': profile.patronymic,
            'date_of_birth': profile.date_of_birth,
            'role': 'сотрудник',
            'faculty': profile.faculty
        })
    elif hasattr(user, 'guest'):
        profile = user.guest
        initial.update({
            'patronymic': profile.patronymic,
            'date_of_birth': profile.date_of_birth,
            'role': 'гость',
        })
    if request.method == 'POST':
        form = ProfileForm(request.POST, user_instance=user)
        if form.is_valid():
            form.save(user)
            return redirect('user_profile', user_id=user.id)
    else:
        form = ProfileForm(user_instance=user, initial=initial)
    return render(request, 'admin/edit_user_profile.html', {
        'form': form,
        'user': user,
    })


@admin_required
# Выгрузка профиля конкретного пользователя в файл
def export_to_excel_user(request):
    if request.method == 'POST' and request.POST.get('export'):
        selected_test = request.POST.get('selected_test', '')
        user_id = request.POST.get('user_id', '')
        if not selected_test or not user_id:
            return HttpResponse("Ошибка: тест или пользователь не выбран", status=400)

        user = get_object_or_404(User, id=user_id)
        profile = None
        if hasattr(user, 'student'):
            profile = user.student
            profile_type = 'student'
        elif hasattr(user, 'staff'):
            profile = user.staff
            profile_type = 'staff'
        elif hasattr(user, 'guest'):
            profile = user.guest
            profile_type = 'guest'
        user_tests = user.usertest_set.filter(test_id=selected_test).order_by('-completed_at')
        test_title = "Не выбрано"
        question_list = []
        try:
            test_obj = Test.objects.get(id=selected_test)
            test_title = test_obj.title
            question_list = list(test_obj.questions.all().order_by('id'))
        except Test.DoesNotExist:
            test_title = "Не найден"
            question_list = []

        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты теста"
        headers = ['Название теста', 'Дата']
        headers.extend([f'Вопрос {i} (баллы)' for i in range(1, len(question_list) + 1)])
        headers.append('Итог')
        headers.append('Результат')
        ws.append(headers)
        for test_result in user_tests:
            row = [
                test_result.test.title,
                test_result.completed_at.strftime("%d.%m.%Y %H:%M")
            ]
            qid2score = {}
            answers = UserAnswer.objects.filter(user_test=test_result)
            for ans in answers:
                qid2score[ans.question_id] = ans.score
            for question in question_list:
                score = qid2score.get(question.id, "Нет ответа")
                row.append(score)
            row.append(test_result.result)
            row.append(test_result.result_description)
            ws.append(row)

        profile_start_col = len(headers) + 2 
        profile_headers = ['Фамилия', 'Имя', 'Отчество', 'Дата рождения', 'Роль']
        if profile_type == 'student':
            profile_headers.extend(['Факультет', 'Направление', 'Профиль', 'Курс', 'Форма обучения', 'Уровень образования'])
        elif profile_type == 'staff':
            profile_headers.append('Факультет')
        for row_num, header in enumerate(profile_headers, start=1):
            ws.cell(row=row_num, column=profile_start_col, value=header)
        if profile:
            profile_row = [
                user.last_name,
                user.first_name,
                getattr(profile, 'patronymic', ''),
                profile.date_of_birth.strftime("%d.%m.%Y") if profile.date_of_birth else '',
                profile.role,
            ]
            if profile_type == 'student':
                profile_row.extend([
                    getattr(profile, 'faculty', ''),
                    getattr(profile, 'direction', ''),
                    getattr(profile, 'profile', ''),
                    getattr(profile, 'grade', ''),
                    getattr(profile, 'study_form', ''),
                    getattr(profile, 'education_level', '')
                ])
            elif profile_type == 'staff':
                profile_row.append(getattr(profile, 'faculty', ''))
            for row_num, value in enumerate(profile_row, start=1):
                ws.cell(row=row_num, column=profile_start_col + 1, value=value)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=user_test_results.xlsx'
        wb.save(response)
        return response
    return HttpResponse("Ошибка: неверный запрос", status=400)


# Удаление пользователя
@admin_required
def delete_user(request, user_id):
    if request.method == 'POST':
        user = get_object_or_404(User, id=user_id)
        user.delete()
        return redirect('admin_users')
    return redirect('user_profile', user_id=user_id)


def pass_test_date_and_status_filters(user_object, last_test, status, test_date_start, test_date_end):
    if test_date_start:
        if not last_test or not last_test.completed_at or last_test.completed_at.date() < datetime.strptime(test_date_start, "%Y-%m-%d").date():
            return False
    if test_date_end:
        if not last_test or not last_test.completed_at or last_test.completed_at.date() > datetime.strptime(test_date_end, "%Y-%m-%d").date():
            return False
    if status:
        result_value = None
        if hasattr(last_test, 'result'):
            result_value = last_test.result
        elif hasattr(last_test, 'result_description'):
            result_value = last_test.result_description
        if not result_value:
            return False
        if str(result_value).strip().lower() != str(status).strip().lower():
            return False
    return True


@admin_required
# Конструктор теста (создание теста)
def constructor(request, test_id=None):
    if test_id:
        test = get_object_or_404(Test, id=test_id)
    else:
        test = None

    if request.method == 'POST':
        test_form = TestForm(request.POST)
        if test_form.is_valid():
            form_data = test_form.cleaned_data
            time_value = form_data['time'] if request.POST.get('timeLimit') != 'none' else None
            start_datetime = test_form.cleaned_data.get('start_datetime')

            test = Test(
                title=form_data['title'],
                description=form_data['description'],
                time=time_value,
                start_datetime=start_datetime
            )
            if request.POST.get('categorySelect') == 'select':
                categories = request.POST.getlist('category')
                test.assigned_categories = ','.join(categories)
            else:
                test.assigned_categories = 'all'
            test.save()

            question_texts = request.POST.getlist('question_text')
            question_types = request.POST.getlist('question_type')

            for index, question_text in enumerate(question_texts):
                if question_text.strip():
                    is_mandatory_value = request.POST.get(f'is_mandatory_{index}') == 'on'
                    question = Question(
                        test=test,
                        question_text=question_text.strip(),
                        question_type=question_types[index] if index < len(question_types) else None,
                        is_mandatory=is_mandatory_value
                    )
                    question.save()
                    if question_types[index] == "choice":
                        option_index = 0
                        while True:
                            choice_text = request.POST.get(f'choice_text_{index}_{option_index}')
                            score = request.POST.get(f'score_{index}_{option_index}')
                            if choice_text is None:
                                break
                            if choice_text.strip() != "" and score is not None and score.strip() != "":
                                try:
                                    score_value = float(score.strip())
                                    choice = Choice(
                                        question=question,
                                        choice_text=choice_text.strip(),
                                        score=score_value
                                    )
                                    choice.save()
                                except ValueError:
                                    pass
                            option_index += 1
                    if question_types[index] == "scale":
                        min_sign = request.POST.get(f'scale_min_sign_{index}', '').strip()
                        max_sign = request.POST.get(f'scale_max_sign_{index}', '').strip()
                        min_score = request.POST.get(f'scale_min_score_{index}', '').strip()
                        max_score = request.POST.get(f'scale_max_score_{index}', '').strip()
                        scale_type = request.POST.get(f'scale_type_{index}', 'default')
                        scale_inverse = request.POST.get(f'scale_inverse_{index}') == 'on'

                        if min_sign and max_sign and max_score:
                            try:
                                scale = Scale(
                                    question=question,
                                    min_sign=min_sign,
                                    max_sign=max_sign,
                                    min_score=int(min_score),
                                    max_score=int(max_score),
                                    type=scale_type,
                                    inverse=scale_inverse  
                                )
                                scale.save()
                            except ValueError:
                                print(f"Failed to convert max_score to int for question {index}")
            return redirect('setup_test_results', test_id=test.id)
    else:
        test_form = TestForm(instance=test)
    questions = test.questions.all() if test else None
    return render(request, 'admin/constructor.html', {
        'test_form': test_form,
        'test': test,
        'questions': test.questions.all() if test else None,
    })


@admin_required
# Создание результатов теста
def setup_test_results(request, test_id):
    test = get_object_or_404(Test, id=test_id)
    if request.method == 'POST':
        results = []
        index = 0
        errors = []
        while True:
            condition = request.POST.get(f'conditions[{index}][condition]', None)
            min_score = request.POST.get(f'conditions[{index}][min_score]', None)
            max_score = request.POST.get(f'conditions[{index}][max_score]', None)
            result_description = request.POST.get(f'conditions[{index}][result_description]', None)
            additional_info = request.POST.get(f'conditions[{index}][additional_info]', '')
            use_image = request.POST.get(f'conditions[{index}][use_image]', 'off') == 'on'
            image = request.FILES.get(f'conditions[{index}][image]', None)
            use_thermometer = request.POST.get(f'conditions[{index}][use_thermometer]', 'off') == 'on'
            color_direction = request.POST.get(f'conditions[{index}][color_direction]', None)
            if condition is None and min_score is None and max_score is None:
                break

            results.append({
                'condition': condition,
                'min_score': min_score,
                'max_score': max_score,
                'result_description': result_description,
                'additional_info': additional_info,
                'use_image': use_image,
                'image': image,
                'use_thermometer': use_thermometer,
                'color_direction': color_direction 
            })
            index += 1
        if not any(ResultData['additional_info'] for ResultData in results):
            errors.append("Пожалуйста, добавьте хотя бы один результат перед созданием теста.")
            return render(request, 'admin/setup_test_results.html', {
                'test': test,
                'results': results,
                'errors': errors,
            })

        for result_data in results:
            min_value = int(result_data['min_score']) if result_data['min_score'] else None
            if result_data['condition'] in ['greater', 'greater_or_equal', 'less', 'less_or_equal', 'equal']:
                max_value = None
            else:
                max_value = int(result_data['max_score']) if result_data['max_score'] else None
            if min_value is not None:
                existing_result = None
                try:
                    existing_result = Result.objects.filter(
                        test=test,
                        min_score=min_value,
                        max_score=max_value,
                        condition=result_data['condition']
                    ).first()
                except Result.DoesNotExist:
                    existing_result = None
                if existing_result:
                    existing_result.result_description = result_data['result_description']
                    existing_result.additional_info = result_data['additional_info']
                    existing_result.use_image = result_data['use_image']
                    if result_data['use_image'] and result_data['image']:
                        existing_result.image = result_data['image']
                    else:
                        existing_result.image = None
                    existing_result.save()
                else:
                    result_kwargs = {
                        'test': test,
                        'min_score': min_value,
                        'max_score': max_value,
                        'result_description': result_data['result_description'],
                        'additional_info': result_data['additional_info'],
                        'condition': result_data['condition'],
                        'use_thermometer': result_data['use_thermometer'],
                        'color_direction': result_data['color_direction'] if result_data['use_thermometer'] else None
                    }
                    if result_data['use_image'] and result_data['image']:
                        result_kwargs['image'] = result_data['image']
                    else:
                        result_kwargs['image'] = None

                    try:
                        Result.objects.create(**result_kwargs)
                    except (ValueError, TypeError) as e:
                        print(f"Ошибка при сохранении результата: {e}")
        if errors:
            return render(request, 'admin/setup_test_results.html', {
                'test': test,
                'results': results,
                'errors': errors,
            })
        return redirect('edit_test_list')
    results = Result.objects.filter(test=test)
    return render(request, 'admin/setup_test_results.html', {
        'test': test,
        'results': results,
    })


@admin_required
# Странице всех тестов для админа
def edit_test_list(request, test_id=None):
    if test_id:
        test = get_object_or_404(Test, id=test_id)
        questions = test.questions.all()
        question_data = [{
            'question': question,
            'choices': question.choices.all(),
            'scale': question.scales.first()
        } for question in questions]
        return render(request, 'admin/constructor.html', {
            'test_form': test, 
            'questions_data': question_data,
        })
    tests = Test.objects.all()
    tests_count = tests.count()
    return render(request, 'admin/edit_test_list.html', {
        'tests': tests,
        'tests_count': tests_count,
    })


@admin_required
# Данные про факультеты и их содержимое
def data_editing(request):
    selected = request.GET.get('level', '')
    levels = [
        ('bachelor_specialist', 'Бакалавриат/Специалитет'),
        ('master', 'Магистратура'),
        ('phd', 'Аспирантура'),
    ]
    faculties = []
    if selected == 'bachelor_specialist':
        faculties_qs = Faculty.objects.order_by('name')
        for faculty in faculties_qs:
            directions = BachSpecDirection.objects.filter(faculty=faculty).order_by('direction_name')
            programs = []
            for direction in directions:
                profiles = BachSpecProfile.objects.filter(direction=direction).order_by('profile')
                programs.append({
                    'direction': direction,
                    'profiles': profiles
                })
            if programs:
                faculties.append({'faculty': faculty, 'programs': programs})
    elif selected == 'master':
        faculties_qs = Faculty.objects.order_by('name')
        for faculty in faculties_qs:
            directions = MasterDirection.objects.filter(faculty=faculty).order_by('direction_name')
            programs = []
            for direction in directions:
                profiles = MasterProfile.objects.filter(direction=direction).order_by('profile')
                programs.append({
                    'direction': direction,
                    'profiles': profiles
                })
            if programs:
                faculties.append({'faculty': faculty, 'programs': programs})
    elif selected == 'phd':
        faculties_qs = Faculty.objects.order_by('name')
        for faculty in faculties_qs:
            directions = PhdDirection.objects.filter(faculty=faculty).order_by('direction_name')
            programs = []
            for direction in directions:
                profiles = PhdProfile.objects.filter(direction=direction).order_by('profile')
                programs.append({
                    'direction': direction,
                    'profiles': profiles
                })
            if programs:
                faculties.append({'faculty': faculty, 'programs': programs})

    return render(request, 'admin/data_editing.html', {
        'levels': levels,
        'selected': selected,
        'faculties': faculties,
        'editable': True,
    })


@admin_required
@csrf_exempt
# Конструктор факультетов
def programs_bulk_update(request):
    level = request.GET.get('level', 'bachelor_specialist')
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Method not allowed"}, status=405)
    try:
        data = json.loads(request.body)
        programs = data.get("programs", {})
        name_changes = data.get("name_changes", [])
        new_faculties = data.get("new_faculties", [])
        new_directions = data.get("new_directions", [])
        new_profiles = data.get("new_profiles", [])
        # Создание новых факультетов
        for faculty_data in new_faculties:
            faculty_name = faculty_data.get("name")
            if faculty_name:
                faculty, created = Faculty.objects.get_or_create(name=faculty_name)
                if created:
                    faculty.save()
                for direction_data in faculty_data.get("directions", []):
                    direction_name = direction_data.get("direction_name")
                    if direction_name:
                        direction_model_key = direction_data.get("direction_model")
                        if direction_model_key == 'bachelor_specialist':
                            direction = BachSpecDirection(faculty=faculty, direction_name=direction_name)
                        elif direction_model_key == 'master':
                            direction = MasterDirection(faculty=faculty, direction_name=direction_name)
                        elif direction_model_key == 'phd':
                            direction = PhdDirection(faculty=faculty, direction_name=direction_name)
                        else:
                            continue
                        direction.save()
                        # Создание профилей для нового направления
                        for profile_data in direction_data.get("profiles", []):
                            profile_name = profile_data.get("profile", "")
                            education_form = profile_data.get("education_form", "")
                            number_of_courses = int(profile_data.get("number_of_courses", 0))
                            if direction_model_key == 'bachelor_specialist':
                                level = profile_data.get("level", "")
                                profile = BachSpecProfile(
                                    direction=direction,
                                    profile=profile_name,
                                    education_form=education_form,
                                    number_of_courses=number_of_courses,
                                    level=level
                                )
                            elif direction_model_key == 'master':
                                profile = MasterProfile(
                                    direction=direction,
                                    profile=profile_name,
                                    education_form=education_form,
                                    number_of_courses=number_of_courses
                                )
                            elif direction_model_key == 'phd':
                                profile = PhdProfile(
                                    direction=direction,
                                    profile=profile_name,
                                    education_form=education_form,
                                    number_of_courses=number_of_courses
                                )
                            profile.save()
        direction_models = {
            'bachelor_specialist': BachSpecDirection,
            'master': MasterDirection,
            'phd': PhdDirection,
        }
        profile_models = {
            'bachelor_specialist': BachSpecProfile,
            'master': MasterProfile,
            'phd': PhdProfile,
        }
        tempid_to_direction_id = {}
        # Удаление факультетов
        deleted_faculties = data.get("deleted_faculties", [])
        for df in deleted_faculties:
            fac_id = df.get("id")
            level = df.get("level")
            if fac_id:
                direction_models = {
                    'bachelor_specialist': BachSpecDirection,
                    'master': MasterDirection,
                    'phd': PhdDirection,
                }
                profile_models = {
                    'bachelor_specialist': BachSpecProfile,
                    'master': MasterProfile,
                    'phd': PhdProfile,
                }
                dir_model = direction_models.get(level)
                prof_model = profile_models.get(level)
                dirs = dir_model.objects.filter(faculty_id=fac_id)
                profs = prof_model.objects.filter(direction__in=dirs)
                profs.delete()
                dirs.delete()
        # Удаление направлений и профилей
        deleted_directions = data.get("deleted_directions", [])
        deleted_profiles = data.get("deleted_profiles", [])
        for dd in deleted_directions:
            dir_id = dd.get("id")
            model = dd.get("model")
            is_new = dd.get("is_new")
            if is_new: continue
            dir_model = direction_models.get(model)
            if dir_model and dir_id:
                dir_model.objects.filter(pk=dir_id).delete()

        for dp in deleted_profiles:
            pid = dp.get("id")
            is_new = dp.get("is_new")
            if is_new: continue
            if not pid: continue
            parts = pid.split(":")
            if len(parts) == 3 and parts[0] == "profile":
                profile_id = parts[1]
                profile_level = parts[2]
                model = profile_models.get(profile_level)
                if model:
                    model.objects.filter(pk=profile_id).delete()
        for nd in new_directions:
            model_key = nd.get("direction_model")
            DirectionModel = direction_models.get(model_key)
            if not DirectionModel: continue
            faculty_id = nd.get("faculty_id")
            direction_name = nd.get("direction_name")
            tempid = nd.get("tempid")
            if not faculty_id or not direction_name or not tempid: continue
            try:
                faculty = Faculty.objects.get(pk=faculty_id)
                direction = DirectionModel(faculty=faculty, direction_name=direction_name)
                direction.save()
                tempid_to_direction_id[tempid] = direction.pk
            except Exception:
                continue

        for change in name_changes:
            direction_model_str = change.get('direction_model')
            direction_model = direction_models.get(direction_model_str)
            direction_id = change.get("direction_id")
            new_name = change.get("new")
            if direction_model and direction_id and new_name:
                direction_model.objects.filter(pk=direction_id).update(direction_name=new_name)

        for pkstr, fields in programs.items():
            tokens = pkstr.split(':')
            if len(tokens) != 3 or tokens[0] != 'profile': continue
            profile_id = tokens[1]
            profile_level = tokens[2]
            ProfileModel = profile_models.get(profile_level)
            if not ProfileModel: continue
            try:
                profile = ProfileModel.objects.get(pk=profile_id)
                for field, value in fields.items():
                    if field == "profile":
                        profile.profile = value
                    elif field == "education_form":
                        profile.education_form = value
                    elif field == "number_of_courses":
                        try:
                            profile.number_of_courses = int(value)
                        except Exception:
                            pass
                    elif field == "level" and hasattr(profile, "level"):
                        profile.level = value
                profile.save()
            except ProfileModel.DoesNotExist:
                continue

        for np in new_profiles:
            parent_direction = np.get("parent_direction")
            profile_level = np.get("profile_level")
            ProfileModel = profile_models.get(profile_level)
            DirectionModel = direction_models.get(profile_level)
            if not ProfileModel or not DirectionModel or not parent_direction:
                continue
            try:
                direction_pk = int(parent_direction)
            except Exception:
                continue 
            try:
                direction = DirectionModel.objects.get(pk=direction_pk)
            except DirectionModel.DoesNotExist:
                continue
            profile_obj = ProfileModel(
                direction=direction,
                profile=np.get("profile", ""),
                education_form=np.get("education_form", ""),
            )
            if 'level' in np and hasattr(profile_obj, 'level'):
                level_value = np.get("level")
                valid_levels = dict(BachSpecProfile.LEVEL_CHOICES)
                if level_value in valid_levels:
                    profile_obj.level = level_value
                else:
                    profile_obj.level = BachSpecProfile.LEVEL_CHOICES[0][0]
            try:
                profile_obj.number_of_courses = int(np.get("number_of_courses") or 0)
            except Exception:
                profile_obj.number_of_courses = 0
            profile_obj.save()
            
        faculty_name_changes = data.get("faculty_name_changes", [])
        for change in faculty_name_changes:
            faculty_id = change.get("faculty_id")
            new_name = change.get("new_name")
            if faculty_id and new_name:
                try:
                    Faculty.objects.filter(pk=faculty_id).update(name=new_name)
                except Exception as e:
                    continue
        return JsonResponse({"success": True})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)


@admin_required
# Редактирование теста
def edit_test(request, test_id):
    from django.db import transaction
    test = get_object_or_404(Test, id=test_id)

    if request.method == 'POST':
        test.title = request.POST.get('title', '')
        test.description = request.POST.get('description', '')
        # Обработка времени
        time_limit = request.POST.get('timeLimit')
        if time_limit == 'none':
            test.time = None
        else:
            time_value = request.POST.get('time', '')
            if time_value.isdigit():
                test.time = int(time_value)
            else:
                test.time = None
        # Обработка времени старта
        start_dt_str = request.POST.get('start_datetime')
        if start_dt_str:
            try:
                test.start_datetime = datetime.strptime(start_dt_str, '%Y-%m-%dT%H:%M')
            except ValueError:
                test.start_datetime = None
        else:
            test.start_datetime = None
        # Обработка назначения теста
        category_select = request.POST.get('categorySelect')
        if category_select == 'assign':
            test.assigned_categories = 'all'
        else:
            categories = request.POST.getlist('category')
            test.assigned_categories = ','.join(categories)
        test.save()
        # Находим все индексы вопросов
        question_indexes = []
        for key in request.POST:
            if key.startswith('question_text_'):
                index = key.split('_')[-1]
                if index.isdigit():
                    question_indexes.append(int(index))
        question_indexes = sorted(set(question_indexes))
        with transaction.atomic():
            seen_question_ids = set()
            for i in question_indexes:
                question_id = request.POST.get(f'question_id_{i}', None)
                question_text = request.POST.get(f'question_text_{i}', '')
                question_type = request.POST.get(f'question_type_{i}', '')
                is_mandatory = request.POST.get(f'is_mandatory_{i}') == 'on'

                if question_id and question_id.strip():
                    # Существующий вопрос
                    question = get_object_or_404(Question, id=question_id)
                    seen_question_ids.add(question.id)
                    if question_text:
                        question.question_text = question_text
                    if question_type:
                        question.question_type = question_type
                    question.is_mandatory = is_mandatory
                    question.save()
                else:
                    # Новый вопрос
                    question = Question.objects.create(
                        test=test,
                        question_text=question_text,
                        question_type=question_type,
                        is_mandatory=is_mandatory
                    )
                    seen_question_ids.add(question.id)

                # Обновление вариантов ответа
                if question.question_type == 'choice':
                    question.choices.all().delete()
                    option_idx = 0
                    while True:
                        choice_text = request.POST.get(f'choice_text_{i}_{option_idx}', None)
                        score = request.POST.get(f'score_{i}_{option_idx}', None)
                        if choice_text is None:
                            break
                        if choice_text.strip() != "" and score is not None and score.strip() != "":
                            try:
                                score_value = float(score.strip())
                                choice = Choice(
                                    question=question,
                                    choice_text=choice_text.strip(),
                                    score=score_value
                                )
                                choice.save()
                            except ValueError:
                                pass
                        option_idx += 1
                elif question.question_type == 'scale':
                    scale = question.scales.first() if question.scales.exists() else None
                    min_score = request.POST.get(f'scale_min_score_{i}', None)
                    max_score = request.POST.get(f'scale_max_score_{i}', None)
                    min_sign = request.POST.get(f'scale_min_sign_{i}', '')
                    max_sign = request.POST.get(f'scale_max_sign_{i}', '')
                    scale_type = request.POST.get(f'scale_type_{i}', 'default')
                    inverse = bool(request.POST.get(f'scale_inverse_{i}'))
                    if scale:
                        if min_score is not None and str(min_score).strip().isdigit():
                            scale.min_score = int(min_score)
                        if max_score is not None and str(max_score).strip().isdigit():
                            scale.max_score = int(max_score)
                        if min_sign is not None:
                            scale.min_sign = min_sign
                        if max_sign is not None:
                            scale.max_sign = max_sign
                        scale.type = scale_type
                        scale.inverse = inverse
                        scale.save()
                    else:
                        if min_score is not None and max_score is not None:
                            Scale.objects.create(
                                question=question,
                                min_score=int(min_score) if str(min_score).strip().isdigit() else 0,
                                max_score=int(max_score) if str(max_score).strip().isdigit() else 0,
                                min_sign=min_sign,
                                max_sign=max_sign,
                                type=scale_type,
                            )
            Question.objects.filter(test=test).exclude(id__in=seen_question_ids).delete()
        return redirect('edit_test_results', test_id=test.id)
    start_datetime_formatted = ''
    if test.start_datetime:
        start_datetime_formatted = test.start_datetime.strftime('%Y-%m-%dT%H:%M')

    time_limit_radio = 'limited' if test.time is not None else 'none'
    time_value = test.time if test.time is not None else ''
    questions = test.questions.all()
    return render(request, 'admin/edit_test.html', {
        'test': test,
        'questions': questions,
        'start_datetime_formatted': start_datetime_formatted,
        'time_limit_radio': time_limit_radio,
        'time_value': time_value,
    })


@admin_required
# Редактирование результатов теста
def edit_test_results(request, test_id):
    test = get_object_or_404(Test, id=test_id)
    results = Result.objects.filter(test=test).order_by('min_score')
    if request.method == 'POST':
        processed_result_ids = set()
        indices = set()
        for key in request.POST.keys():
            if key.startswith('conditions['):
                idx = key.split('[')[1].split(']')[0]
                indices.add(idx)
        for idx in indices:
            result_id = request.POST.get(f'conditions[{idx}][id]', None)
            condition = request.POST.get(f'conditions[{idx}][condition]', '').strip()
            min_score_raw = request.POST.get(f'conditions[{idx}][min_score]')
            max_score_raw = request.POST.get(f'conditions[{idx}][max_score]')
            additional_info = request.POST.get(f'conditions[{idx}][additional_info]', '').strip()
            result_description = request.POST.get(f'conditions[{idx}][result_description]', '').strip()
            image = request.FILES.get(f'conditions[{idx}][image]', None)
            use_image_str = request.POST.get(f'conditions[{idx}][use_image]', None)
            use_thermometer_str = request.POST.get(f'conditions[{idx}][use_thermometer]', None)  # Добавлено
            color_direction = request.POST.get(f'conditions[{idx}][color_direction]', '')  # Добавлено

            try:
                min_score_int = int(min_score_raw) if min_score_raw not in [None, ''] else None
            except:
                min_score_int = None
            try:
                max_score_int = int(max_score_raw) if max_score_raw not in [None, ''] else None
            except:
                max_score_int = None
            if min_score_int is None:
                continue
            if result_id:
                result_obj = Result.objects.filter(id=result_id, test=test).first()
            else:
                result_obj = Result.objects.filter(
                    test=test,
                    condition=condition,
                    min_score=min_score_int,
                    max_score=max_score_int
                ).first()

            if result_obj:
                changed = False
                if result_obj.additional_info != additional_info:
                    result_obj.additional_info = additional_info
                    changed = True
                if result_obj.result_description != result_description:
                    result_obj.result_description = result_description
                    changed = True

                if use_image_str is None:
                    if result_obj.image:
                        result_obj.image.delete(save=False)
                        result_obj.image = None
                        changed = True
                else:
                    if image:
                        result_obj.image = image
                        changed = True

                result_obj.use_thermometer = True if use_thermometer_str else False
                result_obj.color_direction = color_direction if result_obj.use_thermometer else None
                changed = True
                if changed:
                    result_obj.save()
                processed_result_ids.add(result_obj.id)
            else:
                new_result = Result(
                    test=test,
                    condition=condition,
                    min_score=min_score_int,
                    max_score=max_score_int,
                    additional_info=additional_info,
                    result_description=result_description,
                    use_thermometer=True if use_thermometer_str else False,
                    color_direction=color_direction if use_thermometer_str else None
                )
                if use_image_str is None:
                    new_result.image = None
                elif image:
                    new_result.image = image
                new_result.save()
                processed_result_ids.add(new_result.id)
        all_result_ids = set(Result.objects.filter(test=test).values_list('id', flat=True))
        ids_to_delete = all_result_ids - processed_result_ids
        if ids_to_delete:
            Result.objects.filter(id__in=ids_to_delete).delete()
        return redirect('edit_test_list')
    return render(request, 'admin/edit_test_results.html', {
        'test': test,
        'results': results,
    })


@admin_required
# Удаление теста
@csrf_exempt
def delete_test(request, test_id):
    if request.method == 'POST':
        test = get_object_or_404(Test, id=test_id)
        test.delete()
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=400)


@admin_required
# Статистика тестов
def statistics_view(request):
    tests = Test.objects.all()
    test_id = request.GET.get('test_id')
    start_date = request.GET.get('start_date', "")
    end_date = request.GET.get('end_date', "")
    statistics_data = []
    error_message = None

    if test_id and start_date and end_date and start_date != "дд.мм.гггг" and end_date != "дд.мм.гггг":
        try:
            start_date_obj = datetime.strptime(start_date, '%d.%m.%Y')
            end_date_obj = datetime.strptime(end_date, '%d.%m.%Y')
            end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
        except ValueError:
            error_message = "Неверный формат даты. Пожалуйста, введите даты в формате дд.мм.гггг."
            start_date_obj = None
            end_date_obj = None

        if not error_message:
            try:
                query_filter = {
                    'test_id': int(test_id),
                    'completed_at__gte': start_date_obj,
                    'completed_at__lte': end_date_obj
                }
                user_tests = UserTest.objects.filter(**query_filter)
                if user_tests.exists():
                    user_answers = UserAnswer.objects.filter(user_test__in=user_tests)
                    question_results = defaultdict(list)
                    for answer in user_answers:
                        question_number = answer.question.number if hasattr(answer.question, 'number') else answer.question.id
                        question_results[question_number].append(answer.score)
                    for question_number, results in question_results.items():
                        if results:
                            unique_results = set(results)
                            if len(unique_results) == len(results):
                                stat_mode = "—"
                            else:
                                try:
                                    stat_mode = mode(results)
                                except Exception:
                                    stat_mode = "—"
                            stat_variance = variance(results) if len(results) > 1 else 0
                            stat_std_deviation = stdev(results) if len(results) > 1 else 0
                            stat_mean = mean(results)
                            stat_median = median(results)

                            statistics_data.append({
                                'question_number': question_number,
                                'mode': stat_mode,
                                'variance': round(stat_variance, 3),
                                'std_deviation': round(stat_std_deviation, 3),
                                'mean': round(stat_mean, 3),
                                'median': round(stat_median, 3),
                            })
                else:
                    error_message = "Нет доступных данных для выбранного теста в указанный период."
            except Exception:
                error_message = "Ошибка выборки статистики. Проверьте корректность данных."
    else:
        error_message = None
    context = {
        'tests': tests,
        'statistics_data': statistics_data,
        'start_date': start_date if start_date else "дд.мм.гггг",
        'end_date': end_date if end_date else "дд.мм.гггг",
        'error_message': error_message,
        'test_id': test_id 
    }
    return render(request, 'admin/statistics.html', context)


@admin_required
# Выгрузка в файл статистики тестов
def export_statistics_excel(request):
    test_id = request.GET.get('test_id')
    start_date = request.GET.get('start_date', "")
    end_date = request.GET.get('end_date', "")
    statistics_data = []
    if test_id and start_date and end_date and start_date != "дд.мм.гггг" and end_date != "дд.мм.гггг":
        try:
            start_date_obj = datetime.strptime(start_date, '%d.%m.%Y')
            end_date_obj = datetime.strptime(end_date, '%d.%m.%Y')
            end_date_obj = end_date_obj.replace(hour=23, minute=59, second=59)
        except ValueError:
            return HttpResponse("Неверный формат даты. Пожалуйста, введите даты в формате дд.мм.гггг.")
        query_filter = {
            'test_id': int(test_id),
            'completed_at__gte': start_date_obj,
            'completed_at__lte': end_date_obj
        }
        user_tests = UserTest.objects.filter(**query_filter)
        if user_tests.exists():
            user_answers = UserAnswer.objects.filter(user_test__in=user_tests)
            question_results = defaultdict(list)

            for answer in user_answers:
                question_number = answer.question.number if hasattr(answer.question, 'number') else answer.question.id
                question_results[question_number].append(answer.score)
            for question_number, results in question_results.items():
                if results:
                    unique_results = set(results)
                    if len(unique_results) == len(results):
                        stat_mode = "—"
                    else:
                        try:
                            stat_mode = mode(results)
                        except Exception:
                            stat_mode = "—"
                    stat_variance = variance(results) if len(results) > 1 else 0
                    stat_std_deviation = stdev(results) if len(results) > 1 else 0
                    stat_mean = mean(results)
                    stat_median = median(results)
                    statistics_data.append({
                        'question_number': question_number,
                        'mode': stat_mode,
                        'variance': round(stat_variance, 3),
                        'std_deviation': round(stat_std_deviation, 3),
                        'mean': round(stat_mean, 3),
                        'median': round(stat_median, 3),
                    })
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Статистика"
    headers = ['Номер вопроса', 'Мода', 'Дисперсия', 'Стандартное отклонение', 'Среднее арифметическое', 'Медиана']
    worksheet.append(headers)

    question_numbers = [row['question_number'] for row in statistics_data]
    # Сортируем значения
    question_numbers.sort()
    # Заменяем значения в колонке "Номер вопроса" на отсортированные
    for i, row in enumerate(statistics_data):
        row['question_number'] = question_numbers[i]

    # Добавляем данные таблицы в правильном порядке
    for row in statistics_data:
        worksheet.append([
            row['question_number'],
            row['mode'],
            row['variance'],
            row['std_deviation'],
            row['mean'],
            row['median'],
        ])
    # Получаем название теста
    test = Test.objects.get(id=test_id)
    test_title = test.title
    worksheet.insert_cols(len(headers) + 1, 1)
    worksheet.cell(row=1, column=len(headers) + 2, value="Тест:")
    worksheet.cell(row=1, column=len(headers) + 3, value=test_title)
    worksheet.cell(row=2, column=len(headers) + 2, value="Дата тестирования от:")
    worksheet.cell(row=2, column=len(headers) + 3, value=start_date)
    worksheet.cell(row=3, column=len(headers) + 2, value="Дата тестирования до:")
    worksheet.cell(row=3, column=len(headers) + 3, value=end_date)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=статистика.xlsx'
    workbook.save(response)
    return response


@admin_required
#
def user_logout(request):
    if request.user.is_authenticated:
        logout(request)
    return redirect('user_login')